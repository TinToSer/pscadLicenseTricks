## JMS ##
## https://github.com/tintoser
## Developed by TinToSer
## Love to GPT-4

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import json
import datetime
from openpyxl.styles import Font, Border, Side


def write_to_excel(lic_db,excel_file):
    wb = Workbook()
    ws_lic = wb.active
    ws_lic.title = "Users"
    wb.create_sheet(title="Overview")
    ws_overview = wb["Overview"]
    users_header_required = ["License Group","License Feature","License ID","Checked out By","Email","Machine","Expiry","Checkout Expiry","Checkout Period"]
    overview_header_required = ["License Group","License Feature","License Count","License Checkouts"]
    ws_lic.append(users_header_required)
    ws_overview.append(overview_header_required)


    for each_group_data in lic_db["Groups"].values():
        ws_overview.append(list(each_group_data.values())[:-1])
        for each_member in each_group_data["Members"]:
            row = list(each_member.values())
            ws_lic.append(row)
    
    for ws_n in wb.sheetnames:
        ws=wb[ws_n]
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col[0].column_letter].width = max_length + 10
        
        for cell in ws[1]:
            cell.font = Font(bold=True, size=14)
            cell.border = Border(
                            left=Side(style='thick'),
                            right=Side(style='thick'),
                            top=Side(style='thick'),
                            bottom=Side(style='thick')
                        )
        ws.auto_filter.ref = ws.dimensions

    wb.save(f"{datetime.datetime.now().strftime("%d%m%Y_%H%M%S")}_{excel_file}")
        
def get_pscad_license_usage(username, password):
    session = requests.session()
    headers = {
        "Sec-Ch-Ua-Platform": "\"Windows\"",
        "X-Requested-With": "XMLHttpRequest",
        "Accept-Language": "en-US,en;q=0.9",
        "Sec-Ch-Ua": "\"Chromium\";v=\"135\", \"Not-A.Brand\";v=\"8\"",
        "Sec-Ch-Ua-Mobile": "?0",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36",
        "Accept": "*/*",
        "Origin": "https://mycentre.pscad.com",
        "Sec-Fetch-Site": "same-origin",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Dest": "empty",
        "Referer": "",
        "Accept-Encoding": "gzip, deflate, br",
        "Priority": "u=1, i"
    }
    lic_db = {}
    url = "https://mycentre.pscad.com:443/logics/ajax/login/ajx-login-submission.php"
    files = {
        "postUsername": (None, username),
        "postPassword": (None, password)
    }
    response = session.post(url, headers=headers, files=files)
    if response.ok:
        url = "https://mycentre.pscad.com:443/logics/ajax/workgroup/ajx-wrkCard-view.php"
        response = session.get(url)
        json_data = response.json()
        html_data = json_data.get("layout")
        soup = BeautifulSoup(html_data, 'html.parser')

        members_db={}
        workgroups = soup.find_all('div', class_='cardLineItem')
        search_db = {}
        for wg in workgroups:
            name = wg.find('div', class_='columnWorkgroupsName').text.strip()
            creation = wg.find('div', class_='columnWorkgroupsCreation').text.strip()
            expiration = wg.find('div', class_='columnWorkgroupsExpiration').text.strip()
            maintenance = wg.find('div', class_='columnWorkgroupsMaintenanceExpiry').text.strip()
            members = wg.find('div', class_='columnWorkgroupsMembers').text.strip()
            licenses = wg.find('div', class_='columnWorkgroupsLicenses').text.strip()
            join_codes = wg.find('div', class_='columnWorkgroupsJoinCodes').text.strip()
            search_db[name] = {"Id":wg.get("id"),"Created": creation, "Expiration": expiration,
                "Maintenance Expiry": maintenance, "Members_count": members,"Members":[], "Licenses": licenses, "Join Codes": join_codes}
            lic_db=search_db.copy()
        lic_url = "https://mycentre.pscad.com/logics/ajax/workgroupMember/ajx-license-view.php"
        mem_url = "https://mycentre.pscad.com/logics/ajax/workgroupAdmin/ajx-members-view.php"
        for each_pscad in search_db:
            files = {
                "workgroupID": (None, lic_db[each_pscad].get("Id"))
            }
            response = session.post(mem_url, files=files)
            json_data = response.json()
            html_data = json_data.get("membersLayout")
            mem_soup = BeautifulSoup(html_data, 'html.parser')
            members = mem_soup.find_all('div', class_='cardLineOneMember')
            for each_member in members:
                member_name = each_member.find('div', class_='columnMemberName').text.strip()
                member_email = each_member.find('div', class_='columnMemberEmail').text.strip()
                member_username = each_member.find('div', class_='columnMemberUsername').text.strip()
                if member_name in members_db:
                    print(f"Duplicate member name found: '{member_name}' with email '{member_email}' and username '{member_username}'")
                members_db[member_username] = {
                    "Name": member_name,
                    "Email": member_email,
                    "Username": member_username
                }

            response = session.post(lic_url, files=files)
            json_data = response.json()
            html_data = json_data.get("layout")
            soup = BeautifulSoup(html_data, 'html.parser')
            license_groups = soup.find_all('div', class_='cardLineMulti')
            lic_db["Groups"] = {}
            for group in license_groups:
                lic_group = group.find('div', class_='columnLicensesGroupName').text.strip()
                lic_feature = group.find('div', class_='columnLicensesGroupFeatures').text.strip()
                lic_count = group.find('div', class_='columnLicensesGroupCount').text.strip()
                lic_checkouts = group.find('div', class_='columnLicensesGroupCheckout').text.strip()
                each_inst = f"{lic_group} {lic_feature}"
                lic_db["Groups"][each_inst]={
                    "License Group":lic_group,
                    "License feature":lic_feature,
                    "License count":lic_count,
                    "License Checkouts":lic_checkouts,
                    "Members":[]
                }
                detail_items = group.find_all('div', class_='cardLineTwoInfoLicense')
                for detail in detail_items:
                    license_id = detail.find('div', class_='columnLicensesID').text.strip()
                    license_name = detail.find('div', class_='columnLicensesName').text.strip()
                    expiry = detail.find('div', class_='columnLicensesExpiry').text.strip()
                    machine = detail.find('div', class_='columnLicensesMachine').text.strip()
                    checked_out_by = detail.find('div', class_='columnLicensesCheckout').text.strip()
                    checkout_expiry = detail.find('div', class_='columnLicensesCheckoutExpiry').text.strip()
                    checkout_period = detail.find('div', class_='columnLicensesCheckoutPeriod').text.strip()

                    lic_db["Groups"][each_inst]["Members"].append({
                        'License Group': license_name,
                        'License Feature':lic_feature,
                        'License ID': license_id,
                        'Checked out By': checked_out_by,
                        'Email': members_db.get(checked_out_by, {}).get("Email"),
                        'Machine': machine,
                        'Expiry': expiry,
                        'Checkout Expiry': checkout_expiry,
                        'Checkout Period': checkout_period
                    })


    return lic_db

if __name__ == "__main__":
    username = ""
    password = ""
    json_file = "license_users.json"
    excel_file = "license_users.xlsx"
    lic_db = get_pscad_license_usage(username, password)
    print("[*] Since PSCAD License information is name based, so you can ask users to use username in place of their actual Name in profile, it will help you to populate entries for Email")
    write_to_excel(lic_db,excel_file)
    json.dump(lic_db,open(f"{datetime.datetime.now().strftime("%d%m%Y_%H%M%S")}_{json_file}","w+"),indent=4)
    
